import time
from zhipuai import ZhipuAI
import openpyxl
import os

client = ZhipuAI(api_key="5245d351ce18caff927fa8aa1a507612.7Tv0GL4hid61fdBJ")

wb = openpyxl.load_workbook('E:\\CVMstorage\\showcase\\merged_data.xlsx')

log_file_path = "E:\\CVMstorage\\showcase\\AI_error_log.txt"
time_cost_file_path = "E:\\CVMstorage\\showcase\\AI_time_cost.txt"

answer_times = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Processing sheet: {sheet_name}")
    
    folder_path = f"E:\\CVMstorage\\showcase\\{sheet_name}"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    for row_index in range(2, ws.max_row + 1):

        machine_OS = ws.cell(row=row_index, column=8).value
        
        prompt = f"""
                你是一名专业的云计算性能工程师和服务器运维工程师，擅长性能数据分析。以下是一台云服务器在{ws.cell(row=row_index, column=2).value}测试项目，
                {ws.cell(row=row_index, column=3).value}条件下，OS_1(tlinux3.1)的表现显著劣于OS_2(centos7.4)。请基于计算机组成原理、
                操作系统、服务器运维管理等多方面有关云计算和服务器运维方面的知识，分析并解释两者性能差距的原因，提出性能差异在具体实际云计算业务中的影响，
                并输出markdown格式的性能数据分析表单。
                注意，数据来自虚拟化的云服务器，与裸金属服务器的性能表现可能有所不同。原始数据表格如下：

                | 项目       | 值                                             |
                |------------|------------------------------------------------|
                | test_name  | {ws.cell(row=row_index, column=1).value}       |
                | tool_name  | {ws.cell(row=row_index, column=2).value}       |
                | results_key| {ws.cell(row=row_index, column=3).value}       |
                | metric     | {ws.cell(row=row_index, column=4).value}       |
                | OS_1       | {ws.cell(row=row_index, column=5).value}       |
                | OS_2       | {ws.cell(row=row_index, column=6).value}       |
                | 百分比     | {ws.cell(row=row_index, column=7).value}       |
                | 机型/OS    | {machine_OS}                                   |

                要求:
                1. 回答的基本格式如下：
                | 项目       | 值                                             |
                |------------|------------------------------------------------|
                | test_name  | {ws.cell(row=row_index, column=1).value}       |
                | tool_name  | {ws.cell(row=row_index, column=2).value}       |
                | results_key| {ws.cell(row=row_index, column=3).value}       |
                | metric     | {ws.cell(row=row_index, column=4).value}       |
                | OS_1       | {ws.cell(row=row_index, column=5).value}       |
                | OS_2       | {ws.cell(row=row_index, column=6).value}       |
                | 百分比     | {ws.cell(row=row_index, column=7).value}       |
                | 机型/OS    | {machine_OS}                                   |

                性能差异可能导致以下具体实际云计算业务中的问题：
                -xxx
                -xxx
                ... ...

                性能差距原因分析：
                -xxx
                -xxx
                ... ...
                
                2. 指出具体的性能差异和该性能差异可能导致哪些实际应用的问题,尽可能丰富、详细
                3. 分析OS_1(tlinux)性能数据劣于OS_2(centos7.4)的原因，并分点给出具体、专业、深入、详细的解释
                4. 确保你的所有回答是一个完整的markdown表单，包括表格和解释，不要输出其他内容
                5. 回答不允许输出“以下是...”，“以上是...”，“注意：”，“可能是”等总领性、总结性、提示性、猜测性的语句，只输出分析结果

                请严格按照上述要求作答。
                """
        print(f"Sending API request for row {row_index}: {prompt}")
        
        start_time = time.time()
        
        try:
            response = client.chat.completions.create(
                model="glm-4-plus",  # 模型编码需要修改
                messages=[
                    {"role": "user", "content": prompt}
                ],
                stream=False 
            )
            
            answer = response.choices[0].message.content
            
            end_time = time.time()
            answer_time = end_time - start_time
            answer_times.append(answer_time)
            print(f"Answer time for row {row_index}: {answer_time:.2f} seconds")

            markdown_filename = f"{machine_OS}-{row_index}.md"
            markdown_path = os.path.join(folder_path, markdown_filename)
            with open(markdown_path, 'w', encoding='utf-8') as markdown_file:
                markdown_file.write(f"# {machine_OS}\n\n{answer}")
        
        except Exception as e:
            with open(log_file_path, 'a') as log_file:
                log_file.write(f"Failed to get answer for row {row_index} in sheet {sheet_name}. Error: {e}\n")
#写入错误信息
with open(time_cost_file_path, 'w') as time_cost_file:
    total_time = sum(answer_times)
    average_time = total_time / len(answer_times) if answer_times else 0
    time_cost_file.write("Answer Times:\n")
    for i, t in enumerate(answer_times, start=1):
        time_cost_file.write(f"Row {i}: {t:.2f} seconds\n")
    time_cost_file.write(f"Total Time: {total_time:.2f} seconds\n")
    time_cost_file.write(f"Average Time: {average_time:.2f} seconds\n")

print("Done!\n")
